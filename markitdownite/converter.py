from pathlib import Path
from collections import defaultdict

from markitdownite import paths
from markitdownite.charts import render_chart
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell.cell import Cell
from pydantic import BaseModel, ConfigDict
import re


class CellFormatting(BaseModel):
    bold: bool = False
    italic: bool = False
    strike: bool = False

    @classmethod
    def from_cell(cls, cell: Cell) -> "CellFormatting":
        """Extract inline formatting (bold, italic, strikethrough) from a cell."""
        font = cell.font or {}
        return cls(
            bold=bool(font.bold),
            italic=bool(font.italic),
            strike=bool(font.strikethrough),
        )

    def apply_to(self, value: str) -> str:
        """Apply inline Markdown formatting (bold, italic, strikethrough) to a value."""
        if not value:
            return value

        result = value
        if self.strike:
            result = f"~~{result}~~"
        if self.italic:
            result = f"*{result}*"
        if self.bold:
            result = f"**{result}**"

        return result


class CellAnnotation(BaseModel):
    model_config = ConfigDict(frozen=True)

    fg_color: str | None = None
    bg_color: str | None = None
    border: str | None = None
    category: str | None = None

    @classmethod
    def from_cell(cls, cell: Cell) -> "CellAnnotation":
        """Extract formatting annotations (colors, borders) from a cell."""
        font = cell.font or {}
        fill = cell.fill or {}
        border = cell.border or {}

        fg_color = None
        if font.color and hasattr(font.color, "type") and font.color.type == "rgb":
            rgb = font.color.rgb
            if rgb and isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                fg_color = rgb

        bg_color = None
        if fill.start_color and hasattr(fill.start_color, "type") and fill.start_color.type == "rgb":
            rgb = fill.start_color.rgb
            if rgb and isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                bg_color = rgb

        border_style = None
        if border.left and border.left.style:
            border_style = border.left.style

        return cls(
            fg_color=fg_color,
            bg_color=bg_color,
            border=border_style,
        )


class CellMetadata(BaseModel):
    comment: str | None = None
    link: str | None = None

    @classmethod
    def from_cell(cls, cell: Cell) -> "CellAnnotation":
        """Extract formatting annotations (colors, borders) from a cell."""
        font = cell.font or {}
        fill = cell.fill or {}
        border = cell.border or {}

        fg_color = None
        if font.color and hasattr(font.color, "type") and font.color.type == "rgb":
            rgb = font.color.rgb
            if rgb and isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                fg_color = rgb

        bg_color = None
        if fill.start_color and hasattr(fill.start_color, "type") and fill.start_color.type == "rgb":
            rgb = fill.start_color.rgb
            if rgb and isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                bg_color = rgb

        border_style = None
        if border.left and border.left.style:
            border_style = border.left.style

        return cls(
            fg_color=fg_color,
            bg_color=bg_color,
            border=border_style,
        )


class _CellMetadataExtractor:
    @staticmethod
    def from_cell(cell: Cell) -> CellMetadata:
        """Extract metadata (comments, links) from a cell."""
        comment_text = None
        if cell.comment:
            comment_text = str(cell.comment.text).strip() if cell.comment.text else None

        link_url = None
        if cell.hyperlink:
            link_url = cell.hyperlink.target if cell.hyperlink.target else None

        return CellMetadata(comment=comment_text, link=link_url)


def extract_images(workbook, output_dir: Path) -> dict:
    """Extract all embedded raster images from every worksheet.

    Returns a dict mapping sheet_name to list of image indices for that sheet.
    """
    images_dir = paths.images_dir_path(output_dir)
    images_dir.mkdir(parents=True, exist_ok=True)

    img_counter = 0
    sheet_images: dict[str, list[int]] = {}

    for ws in workbook.worksheets:
        if getattr(ws, "_images", []):
            sheet_images[ws.title] = []
        for img in getattr(ws, "_images", []):
            path = paths.image_path(output_dir, img_counter)
            with open(path, "wb") as f:
                f.write(img._data())
            sheet_images[ws.title].append(img_counter)
            img_counter += 1

    return sheet_images


def _merge_cell_ranges(
    annotations: dict[tuple[int, int], CellAnnotation],
) -> list[tuple[str, CellAnnotation]]:
    """Merge adjacent cells with identical annotations into connected components.

    Uses flood-fill to group cells that share the same annotation and are adjacent
    (horizontally or vertically). Each component is represented as its bounding box.

    Returns list of (range_str, annotation) tuples, sorted by position.
    """
    if not annotations:
        return []

    # Group cells by annotation
    cells_by_annotation = defaultdict(set)
    for pos, annotation in annotations.items():
        cells_by_annotation[annotation].add(pos)

    ranges = []
    for annotation, cells in cells_by_annotation.items():
        visited = set()
        for start_pos in cells:
            if start_pos in visited:
                continue

            # Flood-fill to find all connected neighbors
            cluster = {start_pos}
            stack = [start_pos]
            while stack:
                pos = stack.pop()
                row, col = pos
                for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                    neighbor = (row + dr, col + dc)
                    if neighbor in cells and neighbor not in visited:
                        visited.add(neighbor)
                        cluster.add(neighbor)
                        stack.append(neighbor)

            visited.add(start_pos)

            # Find bounding box for this cluster
            min_row = max_row = min_col = max_col = None
            for row, col in cluster:
                if min_row is None:
                    min_row = max_row = row
                    min_col = max_col = col
                else:
                    min_row = min(min_row, row)
                    max_row = max(max_row, row)
                    min_col = min(min_col, col)
                    max_col = max(max_col, col)

            # If cluster fills its bounding box (solid rectangle), use range; otherwise list individual cells
            cluster_area = len(cluster)
            bbox_area = (max_row - min_row + 1) * (max_col - min_col + 1)

            if cluster_area == bbox_area:
                # Solid rectangle, use range
                start_addr = f"{get_column_letter(min_col)}{min_row}"
                if min_row == max_row and min_col == max_col:
                    range_str = start_addr
                else:
                    end_addr = f"{get_column_letter(max_col)}{max_row}"
                    range_str = f"{start_addr}:{end_addr}"
                ranges.append((range_str, annotation))
            else:
                # Has gaps (e.g., cross pattern), list individual cells
                for pos in sorted(cluster):
                    row, col = pos
                    cell_addr = f"{get_column_letter(col)}{row}"
                    ranges.append((cell_addr, annotation))

    # Sort by position
    ranges.sort(key=lambda x: (int(re.search(r'\d+', x[0]).group()), x[0]))
    return ranges


def _parse_cell_range(range_str: str) -> tuple:
    """Parse Excel range string like 'A1:B5' into ((start_col, start_row), (end_col, end_row)).

    Returns 1-indexed row and column numbers to match Excel semantics.
    """
    match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str)
    if not match:
        return None

    start_col_letter, start_row, end_col_letter, end_row = match.groups()
    # column_index_from_string returns 1-indexed; keep it for Excel semantics
    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)
    start_row = int(start_row)
    end_row = int(end_row)

    return ((start_col, start_row), (end_col, end_row))


def _read_sheet_with_unmerged_cells(
    xlsx_path: Path, sheet_name: str, ws
) -> tuple[pd.DataFrame, dict[tuple[int, int], CellAnnotation], dict[tuple[int, int], CellMetadata]]:
    """Read Excel sheet into DataFrame, filling merged cells, and track annotations and metadata.

    openpyxl is used directly instead of pandas to handle merged cells properly.
    Merged cells that are empty (except top-left) are filled with the top-left value.

    Returns: (DataFrame, annotations_dict, metadata_dict) where annotations_dict maps (row, col) to CellAnnotation
    and metadata_dict maps (row, col) to CellMetadata.
    """
    # Phase 1: Extract raw values and cell objects from worksheet.
    # We need cell objects (not just values) to access formatting (fonts, colors, borders).
    # iter_rows(values_only=False) gives us openpyxl Cell objects with full metadata.
    # data stores plain values (for easier processing), cell_objects stores Cell objects (for formatting).
    data = []
    cell_objects = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
        data.append([cell.value for cell in row])
        cell_objects.append(list(row))

    # Edge case: empty sheet
    if not data:
        return pd.DataFrame(), {}, {}

    # Phase 2: Find the data rectangle bounds (first non-empty row and leftmost non-empty column).
    # Sheets may have empty rows/columns before the table starts (e.g., table at D4).
    # We scan to find where the actual data begins, then only process that rectangle.
    first_row_idx = 0
    first_col_idx = 0
    for row_idx, row_list in enumerate(data):
        if any(cell is not None for cell in row_list):
            first_row_idx = row_idx
            for col_idx, cell in enumerate(row_list):
                if cell is not None:
                    first_col_idx = col_idx
                    break
            break

    # Phase 3: Parse and collect all merged cell ranges from the worksheet.
    # For each merge, extract: range boundaries, top-left value, and top-left formatting.
    # Formatting is inherited from the top-left cell to all other cells in the merge.
    merges = []
    for merged_range in ws.merged_cells.ranges:
        parsed = _parse_cell_range(merged_range.coord)
        if parsed:
            (start_col, start_row), (end_col, end_row) = parsed
            top_left_cell = ws.cell(row=start_row, column=start_col)
            top_left_value = top_left_cell.value
            top_left_formatting = CellFormatting.from_cell(top_left_cell)
            merges.append(
                (
                    (start_col, start_row),
                    (end_col, end_row),
                    top_left_value,
                    top_left_formatting,
                )
            )

    # Phase 4: Fill merged cell ranges with the top-left value and formatting.
    # When openpyxl reads a merged range, only the top-left cell has a value; others are None.
    # We replicate the top-left value (with formatting) to all empty cells in the range.
    # This ensures the DataFrame shows the value in all merged cells, not just the top-left.
    for row_idx, row_list in enumerate(data, 1):
        for (
            start_col,
            start_row,
        ), (end_col, end_row), top_left_value, top_left_formatting in merges:
            if start_row <= row_idx <= end_row:
                # Ensure the row is wide enough (openpyxl pads rows with None as needed).
                while len(row_list) < end_col:
                    row_list.append(None)
                # Fill all empty cells in this row's merge range with the top-left value.
                for col in range(start_col, end_col + 1):
                    if row_list[col - 1] is None:
                        formatted_value = top_left_formatting.apply_to(
                            str(top_left_value) if top_left_value is not None else ""
                        )
                        row_list[col - 1] = formatted_value

    # Phase 5: Apply inline formatting and collect annotations and metadata for cells in the data rectangle.
    # Only process cells that are actually in the final output (from first_row_idx/first_col_idx onwards).
    # Inline formatting (bold, italic, strikethrough) is applied directly to cell values.
    # Annotations (colors, borders) and metadata (comments, links) are collected separately.
    # Coordinates are 1-indexed relative to the data rectangle (A1 = 1,1).
    annotations: dict[tuple[int, int], CellAnnotation] = {}
    metadata: dict[tuple[int, int], CellMetadata] = {}
    for row_idx in range(first_row_idx, len(data)):
        row_list = data[row_idx]
        for col_idx in range(first_col_idx, len(row_list)):
            value = row_list[col_idx]
            if row_idx < len(cell_objects) and col_idx < len(cell_objects[row_idx]):
                cell = cell_objects[row_idx][col_idx]
                # Apply inline Markdown formatting (bold, italic, strikethrough) if the cell has it.
                if value is not None:
                    formatting = CellFormatting.from_cell(cell)
                    formatted_value = formatting.apply_to(str(value))
                    row_list[col_idx] = formatted_value

                # Track annotations (colors, borders) for cells that have them.
                annotation = CellAnnotation.from_cell(cell)
                if annotation.fg_color or annotation.bg_color or annotation.border:
                    data_row = row_idx - first_row_idx + 1
                    data_col = col_idx - first_col_idx + 1
                    annotations[(data_row, data_col)] = annotation

                # Track metadata (comments, links) for cells that have them.
                cell_metadata = _CellMetadataExtractor.from_cell(cell)
                if cell_metadata.comment or cell_metadata.link:
                    data_row = row_idx - first_row_idx + 1
                    data_col = col_idx - first_col_idx + 1
                    metadata[(data_row, data_col)] = cell_metadata

    # Phase 6: Extract the final data rectangle and create the DataFrame.
    # header: first row of the data rectangle (becomes column names in the DataFrame).
    # data_rows: all rows after the header (becomes the data in the DataFrame).
    # Slicing with [first_col_idx:] removes any leading empty columns.
    header = data[first_row_idx][first_col_idx:]
    data_rows = [row[first_col_idx:] for row in data[first_row_idx + 1:]]

    # Create and return the DataFrame with the first row as headers.
    return pd.DataFrame(data_rows, columns=header), annotations, metadata

def excel_to_markdown(
    xlsx_path: str | Path,
    output_dir: str | Path,
) -> None:
    """Convert an Excel workbook to a folder with Markdown and assets.

    Args:
        xlsx_path: Path to the input Excel file.
        output_dir: Directory where output will be written. Created if it doesn't exist.

    Creates output_dir with:
        output.md: Markdown file with worksheets as sections. Each section contains:
            - Level-1 heading with worksheet name
            - Level-2 "Table" heading with DataFrame rendered as GH-flavored table
            - Image links for each rendered chart and embedded image
        charts/: Numbered PNG images (0.png, 1.png, ...) for each chart found
        images/: Numbered PNG images (0.png, 1.png, ...) for embedded images

    Charts are rendered via matplotlib. Empty charts are skipped.
    Embedded images are extracted from worksheet cells and included in markdown.
    """
    xlsx_path = Path(xlsx_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    charts_dir = paths.charts_dir_path(output_dir)
    charts_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)

    sheet_images = extract_images(wb, output_dir)

    md_parts: list[str] = []
    chart_counter = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        df, annotations, metadata = _read_sheet_with_unmerged_cells(xlsx_path, sheet_name, ws)

        md_parts.append(f"# {sheet_name}\n")

        md_parts.append("## Table\n")
        md_parts.append(df.to_markdown(index=False))
        md_parts.append("\n")

        if annotations or metadata:
            merged_ranges = _merge_cell_ranges(annotations) if annotations else []
            md_parts.append("### Annotations\n")
            for range_str, annotation in merged_ranges:
                parts = []
                if annotation.fg_color:
                    parts.append(f"fg_color={annotation.fg_color}")
                if annotation.bg_color:
                    parts.append(f"bg_color={annotation.bg_color}")
                if annotation.border:
                    parts.append(f"border={annotation.border}")
                if parts:
                    md_parts.append(f"- {range_str}: {' '.join(parts)}\n")

            # Add metadata for cells (comments and links)
            for (row, col), cell_metadata in sorted(metadata.items()):
                cell_addr = f"{get_column_letter(col)}{row}"
                parts = []
                if cell_metadata.comment:
                    parts.append(f"comment: {cell_metadata.comment}")
                if cell_metadata.link:
                    parts.append(f"link: {cell_metadata.link}")
                if parts:
                    md_parts.append(f"- {cell_addr}: {' '.join(parts)}\n")

            if merged_ranges or metadata:
                md_parts.append("\n")

        for chart in getattr(ws, "_charts", []):
            chart_path = paths.chart_path(output_dir, chart_counter)
            if render_chart(wb, chart, chart_path):
                md_parts.append(f"![Chart]({chart_path})\n")
            chart_counter += 1

        for img_idx in sheet_images.get(sheet_name, []):
            img_path = paths.image_path(output_dir, img_idx)
            md_parts.append(f"![Image]({img_path})\n")

    paths.output_file_path(output_dir).write_text("\n".join(md_parts))
