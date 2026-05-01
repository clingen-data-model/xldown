from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage

from markitdownite.converter import excel_to_markdown
from markitdownite import paths


def test_chart_rendered_as_image(tmp_path: Path):
    # Create a test Excel file with a chart and embedded image
    wb = Workbook()
    ws = wb.active
    ws.title = "Chart"

    # Add sample data
    ws["A1"] = "Category"
    ws["B1"] = "Value"
    ws["A2"] = "A"
    ws["B2"] = 10
    ws["A3"] = "B"
    ws["B3"] = 20

    # Create a bar chart
    chart = BarChart()
    chart.title = "Chart"
    data = Reference(ws, min_col=2, min_row=1, max_row=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=3)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "D2")

    # Add embedded image
    image_path = Path(__file__).parent / "fixtures" / "bottle.png"
    if image_path.exists():
        img = XLImage(str(image_path))
        ws.add_image(img, "A5")

    # Save the test file
    test_excel = tmp_path / "test.xlsx"
    wb.save(test_excel)

    # Test conversion
    output_dir = tmp_path / "output"

    excel_to_markdown(test_excel, output_dir)

    # Verify chart was rendered
    pngs = list(paths.charts_dir_path(output_dir).glob("*.png"))
    assert len(pngs) >= 1

    # Verify image was extracted
    images = list(paths.images_dir_path(output_dir).glob("*.png"))
    assert len(images) >= 1

    content = paths.output_file_path(output_dir).read_text()
    assert "![Chart](" in content
    assert "![Image](" in content
