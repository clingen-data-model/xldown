"""TEMPORARY: render all chart fixtures to ./ci-charts for tolerance calibration."""
from pathlib import Path

import openpyxl

from xldown.converter import render_chart
from tests.test_charts import (
    CHART_CASES,
    CHART_POSITION,
    FACTORY_SHEET_NAME,
    _update_chart_references,
)

out = Path("ci-charts")
out.mkdir(exist_ok=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)
for name, factory in CHART_CASES:
    factory_wb, chart = factory()
    ws = wb.create_sheet(title=name[:31])
    for row in factory_wb.active.iter_rows(values_only=True):
        ws.append(row)
    _update_chart_references(chart, FACTORY_SHEET_NAME, ws.title)
    ws.add_chart(chart, CHART_POSITION)

for i, (name, _) in enumerate(CHART_CASES):
    ws = wb[name[:31]]
    render_chart(wb, ws._charts[0], out / f"{i}.png")
print("wrote", len(CHART_CASES), "charts")
