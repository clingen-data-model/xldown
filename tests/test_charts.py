"""Parametrized rendering tests — one per openpyxl chart class."""

import re
from pathlib import Path

import openpyxl
from openpyxl.chart import (
    AreaChart,
    AreaChart3D,
    BarChart,
    BarChart3D,
    BubbleChart,
    DoughnutChart,
    LineChart,
    LineChart3D,
    PieChart,
    PieChart3D,
    ProjectedPieChart,
    RadarChart,
    Reference,
    ScatterChart,
    Series,
    StockChart,
    SurfaceChart,
    SurfaceChart3D,
)

from markitdownite.converter import render_chart

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# The factory functions all create workbooks with this sheet name. When consolidating
# multiple charts into a single workbook, we rename each sheet to the chart type name
# and must update the chart's internal references to point to the new sheet.
FACTORY_SHEET_NAME = "Data"

# Cell position where we place each chart on its worksheet. Chosen to avoid overlapping
# with the data (which occupies columns A-B). Any valid cell reference would work.
CHART_POSITION = "D2"


def _update_chart_references(chart, old_sheet_name, new_sheet_name):
    """Update all chart references from old_sheet_name to new_sheet_name.

    In openpyxl, chart data references are stored as formulas in the numRef.f property
    (e.g., 'Data'!$B$1:$B$5). When we consolidate charts from separate workbooks into
    a single workbook with renamed sheets, we need to update these formulas to point
    to the new sheet names.
    """
    for series in chart.series:
        for attr in ["val", "cat", "xVal", "yVal", "zVal", "bubbleSize", "identifiers"]:
            ref_obj = getattr(series, attr, None)
            if ref_obj and hasattr(ref_obj, "numRef") and ref_obj.numRef:
                # numRef is openpyxl's internal representation of numeric cell references in charts.
                # When you create a chart like: chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
                # openpyxl stores this as a formula string in the chart's XML like 'Data'!$B$1:$B$5.
                # The numRef.f property contains this formula string. We update it using regex to change the old sheet name to the new sheet name.
                if ref_obj.numRef.f:
                    ref_obj.numRef.f = re.sub(
                        f"'{re.escape(old_sheet_name)}'!",
                        f"'{new_sheet_name}'!",
                        ref_obj.numRef.f,
                    )

        # Update series title references (for legend) - uses strRef instead of numRef
        if hasattr(series, "title") and series.title:
            if hasattr(series.title, "strRef") and series.title.strRef:
                if series.title.strRef.f:
                    series.title.strRef.f = re.sub(
                        f"'{re.escape(old_sheet_name)}'!",
                        f"'{new_sheet_name}'!",
                        series.title.strRef.f,
                    )


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------


def _basic_wb():
    """Workbook with categories in col A and values in col B (rows 1-5)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FACTORY_SHEET_NAME
    for row in [("A", 10), ("B", 20), ("C", 15), ("D", 30), ("E", 25)]:
        ws.append(row)
    return wb


def _multi_series_wb():
    """Workbook with categories in col A and two series in cols B-C (rows 1-6)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FACTORY_SHEET_NAME
    ws.append(["Category", "Series1", "Series2"])
    for row in [
        ("A", 10, 5),
        ("B", 20, 10),
        ("C", 15, 8),
        ("D", 30, 15),
        ("E", 25, 12),
    ]:
        ws.append(row)
    return wb


def _add_standard(chart, wb):
    """Wire val + cat from _basic_wb and return chart."""
    ws = wb.active
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
    chart.set_categories(Reference(ws, min_col=1, min_row=1, max_row=5))
    return chart


# ---------------------------------------------------------------------------
# Factory functions — each returns (workbook, chart)
# ---------------------------------------------------------------------------


def make_bar():
    wb = _basic_wb()
    return wb, _add_standard(BarChart(), wb)


def make_bar_horizontal():
    wb = _basic_wb()
    chart = BarChart()
    chart.type = "bar"
    return wb, _add_standard(chart, wb)


def make_bar_stacked():
    wb = _multi_series_wb()
    ws = wb.active
    chart = BarChart()
    chart.grouping = "stacked"
    chart.overlap = 100
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=6))
    return wb, chart


def make_bar_pct_stacked():
    wb = _multi_series_wb()
    ws = wb.active
    chart = BarChart()
    chart.grouping = "percentStacked"
    chart.overlap = 100
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=6))
    return wb, chart


def make_bar3d():
    wb = _basic_wb()
    return wb, _add_standard(BarChart3D(), wb)


def make_line():
    wb = _basic_wb()
    return wb, _add_standard(LineChart(), wb)


def make_line3d():
    wb = _basic_wb()
    return wb, _add_standard(LineChart3D(), wb)


def make_pie():
    wb = _basic_wb()
    return wb, _add_standard(PieChart(), wb)


def make_pie3d():
    wb = _basic_wb()
    return wb, _add_standard(PieChart3D(), wb)


def make_projected_pie():
    wb = _basic_wb()
    return wb, _add_standard(ProjectedPieChart(), wb)


def make_doughnut():
    wb = _basic_wb()
    return wb, _add_standard(DoughnutChart(), wb)


def make_area():
    wb = _multi_series_wb()
    ws = wb.active
    chart = AreaChart()
    chart.add_data(Reference(ws, min_col=2, max_col=3, min_row=1, max_row=6))
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=6))
    return wb, chart


def make_area_stacked():
    wb = _multi_series_wb()
    ws = wb.active
    chart = AreaChart()
    chart.grouping = "stacked"
    chart.overlap = 100
    data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=6))
    return wb, chart


def make_area3d():
    wb = _basic_wb()
    return wb, _add_standard(AreaChart3D(), wb)


def make_radar():
    wb = _basic_wb()
    return wb, _add_standard(RadarChart(), wb)


def make_radar_filled():
    wb = _basic_wb()
    chart = RadarChart()
    chart.type = "filled"
    return wb, _add_standard(chart, wb)


def make_scatter():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FACTORY_SHEET_NAME
    for row in [(1, 10), (2, 20), (3, 15), (4, 30), (5, 25)]:
        ws.append(row)
    chart = ScatterChart()
    xvalues = Reference(ws, min_col=1, min_row=1, max_row=5)
    yvalues = Reference(ws, min_col=2, min_row=1, max_row=5)
    chart.series.append(Series(yvalues, xvalues, title="S1"))
    return wb, chart


def make_scatter_line():
    wb, chart = make_scatter()
    chart.scatterStyle = "lineMarker"
    return wb, chart


def make_bubble():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FACTORY_SHEET_NAME
    for row in [(1, 10, 5), (2, 20, 10), (3, 15, 3), (4, 30, 8)]:
        ws.append(row)
    chart = BubbleChart()
    xvalues = Reference(ws, min_col=1, min_row=1, max_row=4)
    yvalues = Reference(ws, min_col=2, min_row=1, max_row=4)
    sizes = Reference(ws, min_col=3, min_row=1, max_row=4)
    chart.series.append(Series(yvalues, xvalues, sizes, title="B1"))
    return wb, chart


def make_stock():
    from openpyxl.chart.axis import ChartLines
    from openpyxl.chart.data_source import NumData, NumVal

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FACTORY_SHEET_NAME
    # Stock chart data: Date, Volume, Open, High, Low, Close
    ws.append(["Date", "Volume", "Open", "High", "Low", "Close"])
    rows = [
        ["2024-01", 20000, 100, 110, 90, 105],
        ["2024-02", 15000, 105, 115, 95, 112],
        ["2024-03", 25000, 112, 120, 108, 118],
    ]
    for row in rows:
        ws.append(row)

    # High-low-close chart
    chart = StockChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=5)
    data = Reference(ws, min_col=4, max_col=6, min_row=1, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    for s in chart.series:
        s.graphicalProperties.line.noFill = True

    # Marker for close series
    s.marker.symbol = "dot"
    s.marker.size = 5
    chart.title = "High-low-close"
    chart.hiLowLines = ChartLines()

    # Excel bug workaround: add dummy cache so hiLowLines display
    # Note from bpb: https://openpyxl.readthedocs.io/en/3.1.2/charts/stock.html
    pts = [NumVal(idx=i) for i in range(len(data) - 1)]
    cache = NumData(pt=pts)
    chart.series[-1].val.numRef.numCache = cache

    return wb, chart


def make_surface():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FACTORY_SHEET_NAME
    for row in [[1, 2, 3, 4], [5, 6, 7, 8], [9, 10, 11, 12]]:
        ws.append(row)
    chart = SurfaceChart()
    data = Reference(ws, min_col=1, min_row=1, max_col=4, max_row=3)
    chart.add_data(data)
    return wb, chart


def make_surface3d():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FACTORY_SHEET_NAME
    for row in [[1, 2, 3, 4], [5, 6, 7, 8], [9, 10, 11, 12]]:
        ws.append(row)
    chart = SurfaceChart3D()
    data = Reference(ws, min_col=1, min_row=1, max_col=4, max_row=3)
    chart.add_data(data)
    return wb, chart


# ---------------------------------------------------------------------------
# Parametrized test
# ---------------------------------------------------------------------------

CHART_CASES = [
    ("bar_col", make_bar),
    ("bar_horizontal", make_bar_horizontal),
    ("bar_stacked", make_bar_stacked),
    ("bar_pct_stacked", make_bar_pct_stacked),
    ("bar3d", make_bar3d),
    ("line", make_line),
    ("line3d", make_line3d),
    ("pie", make_pie),
    ("pie3d", make_pie3d),
    ("projected_pie", make_projected_pie),
    ("doughnut", make_doughnut),
    ("area", make_area),
    ("area_stacked", make_area_stacked),
    ("area3d", make_area3d),
    ("radar", make_radar),
    ("radar_filled", make_radar_filled),
    ("scatter", make_scatter),
    ("scatter_line", make_scatter_line),
    ("bubble", make_bubble),
    ("stock", make_stock),
    ("surface", make_surface),
    ("surface3d", make_surface3d),
]


def test_chart_images_match_fixtures(tmp_path: Path):
    """Verify all 22 chart images match their expected fixtures byte-for-byte."""
    fixtures_dir = Path(__file__).parent / "fixtures" / "expected_charts"

    # Create consolidated workbook with all chart types
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, factory in CHART_CASES:
        factory_wb, chart = factory()
        factory_ws = factory_wb.active

        ws = wb.create_sheet(title=name[:31])
        for row in factory_ws.iter_rows(values_only=True):
            ws.append(row)

        _update_chart_references(chart, FACTORY_SHEET_NAME, ws.title)
        ws.add_chart(chart, CHART_POSITION)

    # Render each chart and compare with fixture
    for i, (name, _) in enumerate(CHART_CASES):
        ws = wb[name[:31]]
        chart = ws._charts[0] if ws._charts else None
        assert chart is not None, f"{name}: No chart found"

        output_path = tmp_path / f"{i}.png"
        result = render_chart(wb, chart, output_path)
        assert result, f"{name} ({i}): render_chart returned False"

        with open(output_path, "rb") as f:
            generated = f.read()
        with open(fixtures_dir / f"{i}.png", "rb") as f:
            expected = f.read()

        assert generated == expected, f"{name} ({i}): Generated image differs from fixture"
