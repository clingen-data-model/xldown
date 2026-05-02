from openpyxl import Workbook
from openpyxl.styles import Font
from markitdownite.cells import find_cell_regions, RegionKind, fill_merged_cells, CellAnnotation, group_annotation_ranges


def test_find_cell_regions():
    """Test find_cell_regions: empty data, single cells, tables, and multiple regions."""
    # Empty data
    assert find_cell_regions([]) == []

    # Single cell (PROSE)
    regions = find_cell_regions([["hello"]])
    assert len(regions) == 1
    assert regions[0].kind == RegionKind.PROSE
    assert regions[0].cells == {(0, 0)}

    # 2x2 table (TABLE)
    regions = find_cell_regions([["a", "b"], ["c", "d"]])
    assert len(regions) == 1
    assert regions[0].kind == RegionKind.TABLE
    assert regions[0].cells == {(0, 0), (0, 1), (1, 0), (1, 1)}

    # L-shaped region (TABLE because 3 cells)
    regions = find_cell_regions([["a", "b"], ["c", None]])
    assert len(regions) == 1
    assert regions[0].kind == RegionKind.TABLE
    assert regions[0].cells == {(0, 0), (0, 1), (1, 0)}

    # Multiple disjoint tables, sorted by position
    regions = find_cell_regions([
        ["a", "b", None, "c", "d"],
        ["e", "f", None, "g", "h"]
    ])
    assert len(regions) == 2
    assert regions[0].kind == RegionKind.TABLE
    assert regions[0].cells == {(0, 0), (0, 1), (1, 0), (1, 1)}
    assert regions[1].kind == RegionKind.TABLE
    assert regions[1].cells == {(0, 3), (0, 4), (1, 3), (1, 4)}


def test_fill_merged_cells():
    """Test fill_merged_cells expands Excel merged cells with top-left value and formatting."""
    wb = Workbook()
    ws = wb.active

    # Set up: A1:B2 merged (bold value), D1:D2 merged (italic value)
    ws['A1'] = 'bold_value'
    ws['A1'].font = Font(bold=True)
    ws.merge_cells('A1:B2')

    ws['D1'] = 'italic_value'
    ws['D1'].font = Font(italic=True)
    ws.merge_cells('D1:D2')

    # Build data array matching worksheet structure
    data = [[None, None, None, None, None],
            [None, None, None, None, None]]

    fill_merged_cells(ws, data)

    # A1:B2 should be filled with bold-formatted value
    assert data[0][0] == '**bold_value**'
    assert data[0][1] == '**bold_value**'
    assert data[1][0] == '**bold_value**'
    assert data[1][1] == '**bold_value**'

    # D1:D2 should be filled with italic-formatted value
    assert data[0][3] == '*italic_value*'
    assert data[1][3] == '*italic_value*'


def test_group_annotation_ranges():
    """Test group_annotation_ranges groups adjacent cells with same annotation and handles gaps."""
    # Empty annotations
    assert group_annotation_ranges({}) == []

    # Single cell annotation (1-indexed coordinates)
    annot = CellAnnotation(fg_color="FF0000")
    ranges = group_annotation_ranges({(1, 1): annot})
    assert len(ranges) == 1
    assert ranges[0] == ("A1", annot)

    # Solid rectangle (2x2)
    ranges = group_annotation_ranges({
        (1, 1): annot, (1, 2): annot,
        (2, 1): annot, (2, 2): annot,
    })
    assert len(ranges) == 1
    assert ranges[0] == ("A1:B2", annot)

    # L-shaped cluster (gap at (2, 2)) - should list individual cells
    ranges = group_annotation_ranges({
        (1, 1): annot, (1, 2): annot,
        (2, 1): annot,
    })
    assert len(ranges) == 3
    cell_addresses = [r[0] for r in ranges]
    assert "A1" in cell_addresses
    assert "B1" in cell_addresses
    assert "A2" in cell_addresses

    # Multiple disjoint clusters with different annotations
    annot2 = CellAnnotation(bg_color="00FF00")
    ranges = group_annotation_ranges({
        (1, 1): annot, (1, 2): annot,
        (3, 4): annot2, (3, 5): annot2,
    })
    assert len(ranges) == 2
    # First cluster (solid 1x2)
    assert ranges[0] == ("A1:B1", annot)
    # Second cluster (solid 1x2, sorted after)
    assert ranges[1] == ("D3:E3", annot2)
