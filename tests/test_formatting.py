from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from xldown.converter import excel_to_markdown
from xldown import paths


def test_prose_before_table(tmp_path: Path):
    """Verify converter handles prose text before a table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "WithProse"

    # Add prose before the table
    ws["A1"] = "All values are in TPM units."

    # Add table starting at row 3
    ws["A3"] = "Gene"
    ws["B3"] = "Expression"
    ws["A4"] = "BRCA1"
    ws["B4"] = 1234
    ws["A5"] = "TP53"
    ws["B5"] = 5678

    test_excel = tmp_path / "test.xlsx"
    wb.save(test_excel)

    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    content = paths.output_file_path(output_dir).read_text()
    expected = """# WithProse

All values are in TPM units.

## Table

| Gene   |   Expression |
|:-------|-------------:|
| BRCA1  |         1234 |
| TP53   |         5678 |

"""
    assert content == expected


def test_multiple_tables_same_sheet(tmp_path: Path):
    """Verify converter handles multiple separate tables in one sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MultipleTables"

    # First table (rows 1-3)
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["A2"] = "Alice"
    ws["B2"] = 95
    ws["A3"] = "Bob"
    ws["B3"] = 87

    # Second table (rows 5-7, with gap)
    ws["A5"] = "Product"
    ws["B5"] = "Price"
    ws["A6"] = "Apple"
    ws["B6"] = 1.50
    ws["A7"] = "Banana"
    ws["B7"] = 0.75

    test_excel = tmp_path / "test.xlsx"
    wb.save(test_excel)

    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    content = paths.output_file_path(output_dir).read_text()
    expected = """# MultipleTables

## Table 1

| Name   |   Score |
|:-------|--------:|
| Alice  |      95 |
| Bob    |      87 |

## Table 2

| Product   |   Price |
|:----------|--------:|
| Apple     |    1.5  |
| Banana    |    0.75 |

"""
    assert content == expected


def test_all_formatting_types(tmp_path: Path):
    """Verify inline formatting (bold, italic, strikethrough) and annotations (colors)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Formatting"

    # Headers with bold
    ws["A1"] = "Name"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = "Status"
    ws["B1"].font = Font(bold=True)
    ws["C1"] = "Value"
    ws["C1"].font = Font(bold=True)

    # Bold text
    ws["A2"] = "Alice"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = "Active"
    ws["C2"] = 100

    # Italic text
    ws["A3"] = "Bob"
    ws["A3"].font = Font(italic=True)
    ws["B3"] = "Inactive"
    ws["C3"] = 50

    # Strikethrough text
    ws["A4"] = "Charlie"
    ws["A4"].font = Font(strikethrough=True)
    ws["B4"] = "Pending"
    ws["C4"] = 75

    # Bold + Italic
    ws["A5"] = "Diana"
    ws["A5"].font = Font(bold=True, italic=True)
    ws["B5"] = "Active"
    ws["C5"] = 200

    # Red text (annotation) with comment
    ws["A6"] = "Eve"
    ws["A6"].font = Font(color="FFFF0000")
    ws["A6"].comment = Comment("This record needs review", "Author")
    ws["B6"] = "Error"
    ws["C6"] = 0

    # Frank with background on entire Status column
    ws["A7"] = "Frank"
    ws["B7"] = "Warning"
    ws["B7"].hyperlink = Hyperlink(ref="B7", target="https://example.com/docs")
    ws["C7"] = 25

    # Cell-level subscript
    ws["A8"] = "H2O"
    ws["A8"].font = Font(vertAlign="subscript")
    ws["B8"] = "Normal"
    ws["C8"] = 200

    # Character-level rich text subscript (H₂O with only 2 subscript)
    normal_font = InlineFont()
    subscript_font = InlineFont(vertAlign="subscript")
    h2o_rich = CellRichText(
        TextBlock(normal_font, "H"),
        TextBlock(subscript_font, "2"),
        TextBlock(normal_font, "O")
    )
    ws["A9"] = "H2O"
    ws["B9"] = h2o_rich
    ws["C9"] = 100

    # Rotated text
    ws["A10"] = "Rotated 90°"
    ws["A10"].alignment = Alignment(textRotation=90)
    ws["B10"] = "Normal"
    ws["C10"] = 300

    # Triangle pattern (tests non-rectangular cluster handling)
    ws["A11"] = "Grace"
    ws["B11"] = "OK"
    ws["C11"] = 10
    ws["A12"] = "left"
    ws["B12"] = "apex"
    ws["C12"] = "right"

    # Hide column B (Status) to test hidden column marking
    ws.column_dimensions['B'].hidden = True

    # Apply yellow background to entire Status column (annotation merging test)
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    for row in range(1, 11):
        ws[f"B{row}"].fill = yellow_fill

    # Apply blue background to triangle (annotation gap handling test)
    blue_fill = PatternFill(start_color="FF0000FF", end_color="FF0000FF", fill_type="solid")
    ws["A12"].fill = blue_fill
    ws["B12"].fill = blue_fill
    ws["C12"].fill = blue_fill

    test_excel = tmp_path / "test.xlsx"
    wb.save(test_excel)

    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    content = paths.output_file_path(output_dir).read_text()
    expected = """# Formatting

## Table

| **Name**           | **Status** (hidden)   | **Value**   |
|:-------------------|:----------------------|:------------|
| **Alice**          | Active                | 100         |
| *Bob*              | Inactive              | 50          |
| ~~Charlie~~        | Pending               | 75          |
| ***Diana***        | Active                | 200         |
| Eve                | Error                 | 0           |
| Frank              | Warning               | 25          |
| <sub>H2O</sub>     | Normal                | 200         |
| H2O                | H<sub>2</sub>O        | 100         |
| Rotated 90° (↻90°) | Normal                | 300         |
| Grace              | OK                    | 10          |
| left               | apex                  | right       |

### Annotations
*(Cell references are relative to the table above)*

- B1:B10: bg_color=FFFFFF00

- A6: fg_color=FFFF0000

- A12:C12: bg_color=FF0000FF

- A6: comment: This record needs review

- B7: link: https://example.com/docs

"""
    assert content == expected


def test_isolated_cell_as_prose(tmp_path: Path):
    """Verify a single isolated cell is emitted as prose, not a table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OnlyProse"
    ws["C3"] = "Note: values are provisional"

    test_excel = tmp_path / "test.xlsx"
    wb.save(test_excel)

    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    content = paths.output_file_path(output_dir).read_text()
    expected = """# OnlyProse

Note: values are provisional

"""
    assert content == expected


def test_tables_side_by_side(tmp_path: Path):
    """Verify two tables separated by an empty column produce separate table sections."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SideBySide"

    # First table at A1:B3
    ws["A1"] = "Name"
    ws["B1"] = "Score"
    ws["A2"] = "Alice"
    ws["B2"] = 95
    ws["A3"] = "Bob"
    ws["B3"] = 87

    # Column C is empty (the gap)
    # Second table at D1:E3
    ws["D1"] = "Product"
    ws["E1"] = "Price"
    ws["D2"] = "Apple"
    ws["E2"] = 1.50
    ws["D3"] = "Banana"
    ws["E3"] = 0.75

    test_excel = tmp_path / "test.xlsx"
    wb.save(test_excel)

    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    content = paths.output_file_path(output_dir).read_text()
    expected = """# SideBySide

## Table 1

| Name   |   Score |
|:-------|--------:|
| Alice  |      95 |
| Bob    |      87 |

## Table 2

| Product   |   Price |
|:----------|--------:|
| Apple     |    1.5  |
| Banana    |    0.75 |

"""
    assert content == expected


def test_prose_between_tables(tmp_path: Path):
    """Verify prose cells between tables are emitted as prose (single table region still uses ## Table)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ProseTableProse"

    # Prose at A1
    ws["A1"] = "Introduction text"

    # Table at A3:B5
    ws["A3"] = "Item"
    ws["B3"] = "Count"
    ws["A4"] = "Apple"
    ws["B4"] = 5
    ws["A5"] = "Orange"
    ws["B5"] = 3

    # Prose at A7
    ws["A7"] = "Footer note"

    test_excel = tmp_path / "test.xlsx"
    wb.save(test_excel)

    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    content = paths.output_file_path(output_dir).read_text()
    # Single table → ## Table (not numbered)
    expected = """# ProseTableProse

Introduction text

## Table

| Item   |   Count |
|:-------|--------:|
| Apple  |       5 |
| Orange |       3 |

Footer note

"""
    assert content == expected
