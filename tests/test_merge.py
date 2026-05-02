from pathlib import Path

from openpyxl import Workbook

from xldown.converter import excel_to_markdown
from xldown import paths


def test_merged_cells_horizontal(tmp_path: Path):
    """Verify that horizontally merged header cells fill all columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Create data with merged header
    ws["A1"] = "Merged Header"
    ws["B1"] = "Original B1"  # Will be overwritten by merge
    ws["A2"] = "Value 1"
    ws["B2"] = "Value 2"

    # Merge A1:B1 (spans 2 columns) - B1's value is lost, A1's value fills both
    ws.merge_cells("A1:B1")

    # Save the test file
    test_excel = tmp_path / "test_merged.xlsx"
    wb.save(test_excel)

    # Convert
    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    # Verify the full markdown output
    content = paths.output_file_path(output_dir).read_text()
    expected = """# Data

## Table

| Merged Header   | Merged Header   |
|:----------------|:----------------|
| Value 1         | Value 2         |

"""
    assert content == expected


def test_merged_cells_vertical(tmp_path: Path):
    """Verify that vertically merged cells fill all rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Create data with merged rows
    ws["A1"] = "Header"
    ws["B1"] = "Value Col"
    ws["A2"] = "Merged Value"
    ws["A3"] = None  # Will be filled by merge
    ws["B2"] = "Row 1"
    ws["B3"] = "Row 2"

    # Merge A2:A3 (spans 2 rows)
    ws.merge_cells("A2:A3")

    # Save the test file
    test_excel = tmp_path / "test_merged_v.xlsx"
    wb.save(test_excel)

    # Convert
    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    # Verify the full markdown output
    content = paths.output_file_path(output_dir).read_text()
    expected = """# Data

## Table

| Header       | Value Col   |
|:-------------|:------------|
| Merged Value | Row 1       |
| Merged Value | Row 2       |

"""
    assert content == expected


def test_merged_cells_rectangular(tmp_path: Path):
    """Verify that rectangular merged cells (rows and columns) fill all cells."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Create a 3x3 grid with a rectangular merge A1:B2
    ws["A1"] = "Merged Box"
    ws["B1"] = "Original B1"
    ws["C1"] = "C1"
    ws["A2"] = "Original A2"
    ws["B2"] = "Original B2"
    ws["C2"] = "C2"
    ws["A3"] = "A3"
    ws["B3"] = "B3"
    ws["C3"] = "C3"

    # Merge A1:B2 (2x2 rectangle)
    ws.merge_cells("A1:B2")

    # Save the test file
    test_excel = tmp_path / "test_rect.xlsx"
    wb.save(test_excel)

    # Convert
    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    # Verify the full markdown output
    content = paths.output_file_path(output_dir).read_text()
    expected = """# Data

## Table

| Merged Box   | Merged Box   | C1   |
|:-------------|:-------------|:-----|
| Merged Box   | Merged Box   | C2   |
| A3           | B3           | C3   |

"""
    assert content == expected


def test_multiple_merges(tmp_path: Path):
    """Verify that multiple independent merges in the same sheet work together."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Create data with two separate merges
    ws["A1"] = "Title"
    ws["B1"] = "Subtitle"
    ws["C1"] = "Note"
    ws["A2"] = "Group 1"
    ws["B2"] = "Value 1"
    ws["C2"] = "Value 2"
    ws["A3"] = "Group 2"
    ws["B3"] = "Value 3"
    ws["C3"] = "Value 4"

    # Merge A1:B1 (horizontal)
    ws.merge_cells("A1:B1")
    # Merge A2:A3 (vertical)
    ws.merge_cells("A2:A3")

    # Save the test file
    test_excel = tmp_path / "test_multi.xlsx"
    wb.save(test_excel)

    # Convert
    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    # Verify the full markdown output
    content = paths.output_file_path(output_dir).read_text()
    expected = """# Data

## Table

| Title   | Title   | Note    |
|:--------|:--------|:--------|
| Group 1 | Value 1 | Value 2 |
| Group 1 | Value 3 | Value 4 |

"""
    assert content == expected


def test_table_at_arbitrary_position(tmp_path: Path):
    """Verify that tables can start at any row and column (e.g., D4)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Leave rows 1-3 empty, start table at row 4
    # Leave columns A-C empty, start table at column D
    ws["D4"] = "Name"
    ws["E4"] = "Age"
    ws["D5"] = "Alice"
    ws["E5"] = 30
    ws["D6"] = "Bob"
    ws["E6"] = 25

    test_excel = tmp_path / "test_position.xlsx"
    wb.save(test_excel)

    output_dir = tmp_path / "output"
    excel_to_markdown(test_excel, output_dir)

    # Verify the full markdown output
    content = paths.output_file_path(output_dir).read_text()
    expected = """# Data

## Table

| Name   |   Age |
|:-------|------:|
| Alice  |    30 |
| Bob    |    25 |

"""
    assert content == expected
