from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string, get_column_letter

from xldown import paths
from xldown.cells import (
    CellAnnotation,
    CellFormatting,
    CellMetadata,
    RegionKind,
    fill_merged_cells,
    find_cell_regions,
    group_annotation_ranges,
)
from xldown.charts import render_chart


def read_sheet(
    xlsx_path: Path, sheet_name: str, ws, formula_ws=None
) -> list[tuple[RegionKind, pd.DataFrame | str, dict[tuple[int, int], CellAnnotation], dict[tuple[int, int], CellMetadata], set[str]]]:
    """Read Excel sheet and decompose into tables and prose.

    Extracts cell values and formats them, fills Excel merged cells, detects connected components
    of non-empty cells using Union Find, pads rows to uniform width, and returns per-region content.

    Each connected component is classified as:
    - PROSE: single isolated cell (returned as plain text)
    - TABLE: multi-cell region (first row = headers, returned as DataFrame)

    Args:
        xlsx_path: Path to the Excel file (unused but kept for compatibility)
        sheet_name: Name of the sheet being read
        ws: Worksheet object with data_only=True
        formula_ws: Optional worksheet object with data_only=False for extracting formulas

    Returns list of tuples: (kind, content, annotations, metadata, hidden_columns) where:
    - kind: RegionKind.PROSE or RegionKind.TABLE
    - content: str for prose, pd.DataFrame for table
    - annotations/metadata: dicts with 1-indexed coordinates relative to the region
    - hidden_columns: set of Excel column letters hidden in this region
    """
    # Phase 1: Extract raw values and cell objects from worksheet.
    # Each row's list only contains cells up to the last non-empty cell in that row,
    # so rows may have different lengths and need padding later.
    # Collect formulas keyed by (row_idx, col_idx) if formula_ws is provided.
    data: list[list[str | int | float | bool | None]] = []
    cell_objects: list[list[Cell]] = []
    cell_formulas: dict[tuple[int, int], str] = {}

    if formula_ws:
        rows_iter = zip(ws.iter_rows(values_only=False), formula_ws.iter_rows(values_only=False))
    else:
        rows_iter = ((row, None) for row in ws.iter_rows(values_only=False))

    for row_idx, (row, formula_row) in enumerate(rows_iter):
        data.append([cell.value for cell in row])
        cells = list(row)
        if formula_row:
            for col_idx, (cell, formula_cell) in enumerate(zip(cells, formula_row)):
                if formula_cell.value and isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                    cell_formulas[(row_idx, col_idx)] = formula_cell.value
        cell_objects.append(cells)

    if not data:
        return []

    # Phase 2: Fill Excel merged cells in place (parse ranges and fill with top-left value/formatting)
    fill_merged_cells(ws, data)

    # Phase 3: Detect non-contiguous regions using Union Find.
    regions = find_cell_regions(data)
    if not regions:
        return []

    # Detect hidden columns in the worksheet
    hidden_col_indices = set()
    for col_letter in ws.column_dimensions:
        if ws.column_dimensions[col_letter].hidden:
            hidden_col_indices.add(column_index_from_string(col_letter))

    # Phase 4: Process each region.
    results = []
    for region in regions:
        if region.kind == RegionKind.PROSE:
            # Single isolated cell: extract and emit as plain text
            r, c = next(iter(region.cells))
            results.append((RegionKind.PROSE, data[r][c], {}, {}, set()))
            continue

        # Table region: apply formatting, collect annotations/metadata, create DataFrame
        # Pad rows in this region to the region's bounding box width
        for row_idx in range(region.min_row, region.max_row + 1):
            row = data[row_idx]
            width = region.max_col + 1
            row.extend([None] * (width - len(row)))

        annotations: dict[tuple[int, int], CellAnnotation] = {}
        metadata: dict[tuple[int, int], CellMetadata] = {}
        header = None
        data_rows = []

        # Apply inline formatting and collect annotations/metadata
        for row_idx in range(region.min_row, region.max_row + 1):
            row_list = data[row_idx]
            for col_idx in range(region.min_col, region.max_col + 1):
                if col_idx < len(row_list):
                    value = row_list[col_idx]
                    if row_idx < len(cell_objects) and col_idx < len(cell_objects[row_idx]):
                        cell = cell_objects[row_idx][col_idx]
                        # Apply inline Markdown formatting
                        if value is not None:
                            # Check for character-level rich text formatting first (e.g., H₂O)
                            rich_text = CellFormatting.format_rich_text(cell)
                            if rich_text is not None:
                                row_list[col_idx] = rich_text
                            else:
                                # Fall back to cell-level formatting
                                formatting = CellFormatting.from_cell(cell)
                                row_list[col_idx] = formatting.apply_to(str(value))

                        # Track annotations and metadata
                        data_row = row_idx - region.min_row + 1
                        data_col = col_idx - region.min_col + 1

                        annotation = CellAnnotation.from_cell(cell, formula=cell_formulas.get((row_idx, col_idx)))
                        if annotation.fg_color or annotation.bg_color or annotation.border or annotation.formula:
                            annotations[(data_row, data_col)] = annotation

                        cell_metadata = CellMetadata.from_cell(cell)
                        if cell_metadata.comment or cell_metadata.link:
                            metadata[(data_row, data_col)] = cell_metadata

            # Build return data structure
            row_slice = row_list[region.min_col : region.max_col + 1]
            if row_idx == region.min_row:
                header = row_slice
            else:
                data_rows.append(row_slice)

        # Determine which columns in this region are hidden and mark them in header
        region_hidden_cols = set()
        for i, col_idx in enumerate(range(region.min_col + 1, region.max_col + 2)):
            if col_idx in hidden_col_indices:
                region_hidden_cols.add(get_column_letter(col_idx))
                if i < len(header):
                    header[i] = f"{header[i]} (hidden)"

        df = pd.DataFrame(data_rows, columns=header)
        results.append((RegionKind.TABLE, df, annotations, metadata, region_hidden_cols))

    return results

def excel_to_markdown(
    xlsx_path: str | Path,
    output_dir: str | Path,
) -> None:
    """Convert an Excel workbook to a folder with Markdown and assets.

    Args:
        xlsx_path: Path to the input Excel file.
        output_dir: Directory where output will be written. Created if it doesn't exist.

    Creates output_dir with:
        output.md: Markdown file with worksheets as sections. Each section contains:
            - Level-1 heading with worksheet name
            - Level-2 "Table" heading with DataFrame rendered as GH-flavored table
            - Image links for each rendered chart and embedded image
        charts/: Numbered PNG images (0.png, 1.png, ...) for each chart found
        images/: Numbered PNG images (0.png, 1.png, ...) for embedded images

    Charts are rendered via matplotlib. Empty charts are skipped.
    Embedded images are extracted from worksheet cells and included in markdown.
    """
    xlsx_path = Path(xlsx_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    charts_dir = paths.charts_dir_path(output_dir)
    charts_dir.mkdir(parents=True, exist_ok=True)

    images_dir = paths.images_dir_path(output_dir)
    images_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True, rich_text=True)
    wb_formulas = load_workbook(xlsx_path, data_only=False, rich_text=True)

    md_parts: list[str] = []
    chart_counter = 0
    img_counter = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        formula_ws = wb_formulas[sheet_name]
        regions = read_sheet(xlsx_path, sheet_name, ws, formula_ws=formula_ws)

        md_parts.append(f"# {sheet_name}\n")

        # Precompute number of table regions for heading logic
        tables_in_sheet = sum(1 for kind, *_ in regions if kind == RegionKind.TABLE)
        table_counter = 0

        # Process each region (prose or table)
        for kind, content, annotations, metadata, hidden_columns in regions:
            if kind == RegionKind.PROSE:
                md_parts.append(f"{content}\n")
                continue

            # Table region
            table_counter += 1
            heading = f"## Table {table_counter}" if tables_in_sheet > 1 else "## Table"
            md_parts.append(f"{heading}\n")
            md_parts.append(f"{content.to_markdown(index=False)}\n")

            if annotations or metadata:
                merged_ranges = group_annotation_ranges(annotations) if annotations else []
                md_parts.append("### Annotations")
                md_parts.append("*(Cell references are relative to the table above)*\n")
                for range_str, annotation in merged_ranges:
                    parts = []
                    if annotation.fg_color:
                        parts.append(f"fg_color={annotation.fg_color}")
                    if annotation.bg_color:
                        parts.append(f"bg_color={annotation.bg_color}")
                    if annotation.border:
                        parts.append(f"border={annotation.border}")
                    if annotation.formula:
                        parts.append(f"formula={annotation.formula}")
                    if parts:
                        md_parts.append(f"- {range_str}: {' '.join(parts)}\n")

                # Add metadata for cells (comments and links)
                for (row, col), cell_metadata in sorted(metadata.items()):
                    cell_addr = f"{get_column_letter(col)}{row}"
                    parts = []
                    if cell_metadata.comment:
                        parts.append(f"comment: {cell_metadata.comment}")
                    if cell_metadata.link:
                        parts.append(f"link: {cell_metadata.link}")
                    if parts:
                        md_parts.append(f"- {cell_addr}: {' '.join(parts)}\n")

                if merged_ranges or metadata:
                    md_parts.append("\n")

        for chart in getattr(ws, "_charts", []):
            chart_path = paths.chart_path(output_dir, chart_counter)
            if render_chart(wb, chart, chart_path):
                md_parts.append(f"![Chart]({chart_path})\n")
            chart_counter += 1

        for img in getattr(ws, "_images", []):
            img_path = paths.image_path(output_dir, img_counter)
            with open(img_path, "wb") as f:
                f.write(img._data())
            md_parts.append(f"![Image]({img_path})\n")
            img_counter += 1

    output_text = "\n".join(md_parts)
    if output_text:
        output_text = output_text.rstrip('\n') + '\n\n'
    paths.output_file_path(output_dir).write_text(output_text)
