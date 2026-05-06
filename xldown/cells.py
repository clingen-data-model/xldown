import re
from collections import defaultdict
from enum import StrEnum

from openpyxl.cell.cell import Cell
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import BaseModel, ConfigDict


class RegionKind(StrEnum):
    PROSE = "prose"
    TABLE = "table"


class CellFormatting(BaseModel):
    model_config = ConfigDict(frozen=True)

    bold: bool = False
    italic: bool = False
    strike: bool = False
    superscript: bool = False
    subscript: bool = False
    rotation: int | None = None

    @classmethod
    def from_cell(cls, cell: Cell) -> "CellFormatting":
        """Extract inline formatting (bold, italic, strikethrough, superscript, subscript) and rotation from a cell."""
        font = cell.font or {}
        vert_align = getattr(font, 'vertAlign', None)

        rotation = None
        if cell.alignment and cell.alignment.textRotation:
            rotation = cell.alignment.textRotation

        return cls(
            bold=bool(font.bold),
            italic=bool(font.italic),
            strike=bool(font.strikethrough),
            superscript=vert_align == "superscript",
            subscript=vert_align == "subscript",
            rotation=rotation,
        )

    @staticmethod
    def format_rich_text(cell: Cell) -> str:
        """Format rich text cell with character-level subscript/superscript formatting.

        If cell contains rich text with per-character formatting (e.g., H₂O where only
        the "2" is subscript), apply HTML tags to individual text blocks. Otherwise
        returns None to indicate no rich text formatting.
        """
        if isinstance(cell.value, CellRichText):
            result = ""
            for block in cell.value:
                text = block.text
                if block.font and block.font.vertAlign == "subscript":
                    result += f"<sub>{text}</sub>"
                elif block.font and block.font.vertAlign == "superscript":
                    result += f"<sup>{text}</sup>"
                else:
                    result += text
            return result
        return None

    def apply_to(self, value: str) -> str:
        """Apply inline Markdown formatting (bold, italic, strikethrough, superscript, subscript) and rotation marker to a value."""
        if not value:
            return value

        result = value
        if self.strike:
            result = f"~~{result}~~"
        if self.superscript:
            result = f"<sup>{result}</sup>"
        if self.subscript:
            result = f"<sub>{result}</sub>"
        if self.italic:
            result = f"*{result}*"
        if self.bold:
            result = f"**{result}**"

        if self.rotation:
            result = f"{result} (↻{self.rotation}°)"

        return result


class CellAnnotation(BaseModel):
    model_config = ConfigDict(frozen=True)

    fg_color: str | None = None
    bg_color: str | None = None
    border: str | None = None
    formula: str | None = None
    category: str | None = None

    @classmethod
    def from_cell(cls, cell: Cell, formula: str | None = None) -> "CellAnnotation":
        """Extract formatting annotations (colors, borders, formula) from a cell."""
        font = cell.font or {}
        fill = cell.fill or {}
        border = cell.border or {}

        fg_color = None
        if font.color and hasattr(font.color, "type") and font.color.type == "rgb":
            rgb = font.color.rgb
            if rgb and isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                fg_color = rgb

        bg_color = None
        if fill.start_color and hasattr(fill.start_color, "type") and fill.start_color.type == "rgb":
            rgb = fill.start_color.rgb
            if rgb and isinstance(rgb, str) and rgb not in ("00000000", "FFFFFFFF"):
                bg_color = rgb

        border_style = None
        if border.left and border.left.style:
            border_style = border.left.style

        return cls(
            fg_color=fg_color,
            bg_color=bg_color,
            border=border_style,
            formula=formula,
        )


class CellMetadata(BaseModel):
    model_config = ConfigDict(frozen=True)

    comment: str | None = None
    link: str | None = None

    @classmethod
    def from_cell(cls, cell: Cell) -> "CellMetadata":
        """Extract metadata (comments, links) from a cell."""
        comment_text = None
        if cell.comment:
            comment_text = str(cell.comment.text).strip() if cell.comment.text else None

        link_url = None
        if cell.hyperlink:
            link_url = cell.hyperlink.target if cell.hyperlink.target else None

        return cls(comment=comment_text, link=link_url)


class SheetRegion(BaseModel):
    """A connected component of non-empty cells on a sheet.

    All coordinates are 0-indexed into the data list.
    kind is PROSE for single isolated cells, TABLE for multi-cell regions.
    """
    model_config = ConfigDict(frozen=True)

    kind: RegionKind
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    cells: set[tuple[int, int]]


def find_cell_regions(data: list[list]) -> list[SheetRegion]:
    """Detect connected components of non-empty cells using flood-fill.

    Each region is a group of adjacent non-empty cells (4-connected: horizontal/vertical).
    Single isolated cells are marked as PROSE; multi-cell regions are marked as TABLE.

    Returns a list of SheetRegion objects sorted by (min_row, min_col).
    """
    occupied = {(r, c) for r, row in enumerate(data) for c, val in enumerate(row) if val is not None}
    if not occupied:
        return []

    visited = set()
    regions = []

    for start_pos in occupied:
        if start_pos in visited:
            continue

        # Flood-fill to find all connected neighbors
        cluster = {start_pos}
        stack = [start_pos]
        while stack:
            pos = stack.pop()
            row, col = pos
            for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                neighbor = (row + dr, col + dc)
                if neighbor in occupied and neighbor not in visited:
                    visited.add(neighbor)
                    cluster.add(neighbor)
                    stack.append(neighbor)

        visited.add(start_pos)

        # Find bounding box for this cluster
        min_row = min(r for r, c in cluster)
        max_row = max(r for r, c in cluster)
        min_col = min(c for r, c in cluster)
        max_col = max(c for r, c in cluster)
        kind = RegionKind.PROSE if len(cluster) == 1 else RegionKind.TABLE
        regions.append(SheetRegion(
            kind=kind,
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            cells=cluster,
        ))

    # Sort by position (top-to-bottom, left-to-right)
    regions.sort(key=lambda r: (r.min_row, r.min_col))
    return regions


def parse_cell_range(range_str: str) -> tuple:
    """Parse Excel range string like 'A1:B5' into ((start_col, start_row), (end_col, end_row)).

    Returns 1-indexed row and column numbers to match Excel semantics.
    """
    match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", range_str)
    if not match:
        return None

    start_col_letter, start_row, end_col_letter, end_row = match.groups()
    # column_index_from_string returns 1-indexed; keep it for Excel semantics
    start_col = column_index_from_string(start_col_letter)
    end_col = column_index_from_string(end_col_letter)
    start_row = int(start_row)
    end_row = int(end_row)

    return ((start_col, start_row), (end_col, end_row))


def fill_merged_cells(ws, data: list[list]) -> None:
    """Fill Excel merged cell ranges with top-left value and formatting."""
    merges = []
    for merged_range in ws.merged_cells.ranges:
        parsed = parse_cell_range(merged_range.coord)
        if parsed:
            (start_col, start_row), (end_col, end_row) = parsed
            top_left_cell = ws.cell(row=start_row, column=start_col)
            top_left_value = top_left_cell.value
            top_left_formatting = CellFormatting.from_cell(top_left_cell)
            merges.append(
                (
                    (start_col, start_row),
                    (end_col, end_row),
                    top_left_value,
                    top_left_formatting,
                )
            )

    for row_idx, row_list in enumerate(data, 1):
        for (
            start_col,
            start_row,
        ), (end_col, end_row), top_left_value, top_left_formatting in merges:
            if start_row <= row_idx <= end_row:
                while len(row_list) < end_col:
                    row_list.append(None)
                for col in range(start_col, end_col + 1):
                    if row_list[col - 1] is None:
                        formatted_value = top_left_formatting.apply_to(
                            str(top_left_value) if top_left_value is not None else ""
                        )
                        row_list[col - 1] = formatted_value


def group_annotation_ranges(
    annotations: dict[tuple[int, int], CellAnnotation],
) -> list[tuple[str, CellAnnotation]]:
    """Group adjacent cells with identical annotations into connected components.

    Uses flood-fill to group cells that share the same annotation and are adjacent
    (horizontally or vertically). Each component is represented as its bounding box.

    Returns list of (range_str, annotation) tuples, sorted by position.
    """
    if not annotations:
        return []

    # Group cells by annotation
    cells_by_annotation = defaultdict(set)
    for pos, annotation in annotations.items():
        cells_by_annotation[annotation].add(pos)

    ranges = []
    for annotation, cells in cells_by_annotation.items():
        visited = set()
        for start_pos in cells:
            if start_pos in visited:
                continue

            # Flood-fill to find all connected neighbors
            cluster = {start_pos}
            stack = [start_pos]
            while stack:
                pos = stack.pop()
                row, col = pos
                for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                    neighbor = (row + dr, col + dc)
                    if neighbor in cells and neighbor not in visited:
                        visited.add(neighbor)
                        cluster.add(neighbor)
                        stack.append(neighbor)

            visited.add(start_pos)

            # Find bounding box for this cluster
            min_row = max_row = min_col = max_col = None
            for row, col in cluster:
                if min_row is None:
                    min_row = max_row = row
                    min_col = max_col = col
                else:
                    min_row = min(min_row, row)
                    max_row = max(max_row, row)
                    min_col = min(min_col, col)
                    max_col = max(max_col, col)

            # If cluster fills its bounding box (solid rectangle), use range; otherwise list individual cells
            cluster_area = len(cluster)
            bbox_area = (max_row - min_row + 1) * (max_col - min_col + 1)

            if cluster_area == bbox_area:
                # Solid rectangle, use range
                start_addr = f"{get_column_letter(min_col)}{min_row}"
                if min_row == max_row and min_col == max_col:
                    range_str = start_addr
                else:
                    end_addr = f"{get_column_letter(max_col)}{max_row}"
                    range_str = f"{start_addr}:{end_addr}"
                ranges.append((range_str, annotation))
            else:
                # Has gaps (e.g., cross pattern), list individual cells
                for pos in sorted(cluster):
                    row, col = pos
                    cell_addr = f"{get_column_letter(col)}{row}"
                    ranges.append((cell_addr, annotation))

    # Sort by position
    ranges.sort(key=lambda x: (int(re.search(r'\d+', x[0]).group()), x[0]))
    return ranges
